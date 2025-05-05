import pandas as pd
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Min
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from .models import Paciente, Pergunta, Resposta, Alternativa
from .models import FASES_TRATAMENTO, TIPO_PERGUNTA_CHOICES



def index(request):
    return render(request, 'galeria/index.html')

def cadastro_paciente(request):
    if request.method == "POST":
        numero_paciente = request.POST.get("numero_paciente")
        numero_centro = request.POST.get("numero_centro")
        numero_pesquisador = request.POST.get("numero_pesquisador")
        data_observacao = request.POST.get("data_observacao")
        fase_tratamento = request.POST.get("fase_tratamento")

        # Criar o paciente no banco de dados
        paciente = Paciente.objects.create(
            numero_paciente=numero_paciente,
            numero_centro=numero_centro,
            numero_pesquisador=numero_pesquisador,
            data_observacao=data_observacao,
            fase_tratamento=fase_tratamento
        )

        # Obter a primeira pergunta da fase selecionada
        primeira_pergunta = Pergunta.objects.filter(
            numero_pergunta=1,
            fase_tratamento=paciente.fase_tratamento
        ).first()

        if primeira_pergunta:
            return redirect("responder_pergunta", paciente_id=paciente.id, pergunta_id=primeira_pergunta.id)
        else:
            return HttpResponse("Nenhuma pergunta encontrada para a fase selecionada.", status=404)

    return render(request, "galeria/cadastro_paciente.html")


def responder_pergunta(request, paciente_id, pergunta_id=None):
    paciente = get_object_or_404(Paciente, id=paciente_id)

    if pergunta_id:
        pergunta = get_object_or_404(Pergunta, id=pergunta_id)
    else:
        pergunta = Pergunta.objects.filter(numero_pergunta=1, fase_tratamento=paciente.fase_tratamento).first()

    if not pergunta:
        return HttpResponse("Pergunta não encontrada.", status=404)

    if request.method == "POST":
        resposta_raw = request.POST.get("resposta", "").strip().lower()

        # Trata tipo booleano somente se sim/não
        if pergunta.tipo == "sim_nao":
            resposta_convertida = resposta_raw == "sim"
        else:
            resposta_convertida = resposta_raw  # salva como string mesmo

        # Salva resposta
        Resposta.objects.update_or_create(
            paciente=paciente,
            pergunta=pergunta,
            defaults={"resposta": resposta_convertida}
        )

        # Inicializa a pilha se necessário
        if "retornar_para_pilha" not in request.session:
            request.session["retornar_para_pilha"] = []

        # Se for desvio, empilha o retorno correto
        if resposta_raw == "sim" and pergunta.desvio_para:
            # se houver retornar_para, ele é mais confiável do que proxima_se_sim
            if pergunta.retornar_para:
                request.session["retornar_para_pilha"].append(pergunta.retornar_para.id)
            elif pergunta.proxima_se_sim:
                request.session["retornar_para_pilha"].append(pergunta.proxima_se_sim.id)
            request.session.modified = True
            return redirect("responder_pergunta", paciente_id=paciente.id, pergunta_id=pergunta.desvio_para.id)

        # Se chegou ao fim do desvio (sem próxima pergunta)
        if not pergunta.proxima_se_sim and not pergunta.proxima_se_nao:
            if request.session.get("retornar_para_pilha"):
                proxima_id = request.session["retornar_para_pilha"].pop()
                request.session.modified = True
                return redirect("responder_pergunta", paciente_id=paciente.id, pergunta_id=proxima_id)
            else:
                return render(request, "galeria/confirmacao_conclusao.html", {"paciente": paciente})

        # Continua normalmente
        proxima = pergunta.proxima_se_sim if resposta_raw == "sim" else pergunta.proxima_se_nao
        if proxima:
            return redirect("responder_pergunta", paciente_id=paciente.id, pergunta_id=proxima.id)

        return render(request, "galeria/confirmacao_conclusao.html", {"paciente": paciente})

    return render(request, "galeria/questionario.html", {
        "paciente": paciente,
        "pergunta": pergunta
    })


FASES_PRINCIPAIS = ['fase 1', 'fase1b', 'fase 2', 'fase 3', 'redução']

def exportar_relatorio_excel(request):
    # Agrupar respostas para estatísticas, eliminando duplicidades
    respostas_agrupadas = Resposta.objects.values(
        'pergunta__fase_tratamento',
        'pergunta__numero_pergunta',
        'pergunta__texto',
        'resposta'
    ).annotate(total=Count('id'))

    # Contar número de pacientes que responderam cada pergunta (eliminar duplicidades)
    pacientes_por_pergunta = Resposta.objects.values(
        'pergunta__fase_tratamento',
        'pergunta__numero_pergunta'
    ).annotate(total_pacientes=Count('paciente', distinct=True))

    pacientes_dict = {
        (p['pergunta__fase_tratamento'], p['pergunta__numero_pergunta']): p['total_pacientes']
        for p in pacientes_por_pergunta
    }

    # Organizar estatísticas por pergunta
    estatisticas = {}
    for item in respostas_agrupadas:
        chave = (item['pergunta__fase_tratamento'], item['pergunta__numero_pergunta'])
        if chave not in estatisticas:
            estatisticas[chave] = {
                "fase": item['pergunta__fase_tratamento'],
                "numero_pergunta": item['pergunta__numero_pergunta'],
                "pergunta": item['pergunta__texto'],
                "contagem": {},
                "respondentes": pacientes_dict.get(chave, 0)
            }
        estatisticas[chave]["contagem"][item['resposta']] = item['total']

    # Transformar as estatísticas em DataFrame
    dados_estatisticas = []
    for dados in estatisticas.values():
        total = sum(dados['contagem'].values())
        for resposta, qtd in dados['contagem'].items():
            percentual = round((qtd / total) * 100, 2) if total else 0
            dados_estatisticas.append({
                "Fase": dados["fase"],
                "Número Pergunta": dados["numero_pergunta"],
                "Pergunta": dados["pergunta"],
                "Resposta": resposta,
                "Total": qtd,
                "%": percentual,
                "Respondentes": dados["respondentes"]
            })
    df_estatisticas = pd.DataFrame(dados_estatisticas)

    # Dados brutos, codificando em 0 e 1 e eliminando duplicidades
    respostas = Resposta.objects.select_related('paciente', 'pergunta').distinct()
    brutos = []
    for r in respostas:
        cod = 1 if r.resposta.lower() == "sim" else 0 if r.resposta.lower() == "não" else r.resposta
        brutos.append({
            "Paciente": r.paciente.numero_paciente,
            "Centro": r.paciente.numero_centro,
            "Pesquisador": r.paciente.numero_pesquisador,
            "Data da Observação": r.paciente.data_observacao,
            "Fase do Tratamento": r.paciente.fase_tratamento,
            "Número da Pergunta": r.pergunta.numero_pergunta,
            "Texto da Pergunta": r.pergunta.texto,
            "Tipo de Pergunta": r.pergunta.tipo,
            "Resposta": r.resposta,
            "Resposta Codificada": cod,
        })
    df_brutos = pd.DataFrame(brutos)

    # Dicionário de fases
    df_fases = pd.DataFrame([{"Código": k, "Descrição": v} for k, v in FASES_TRATAMENTO])

    # Dicionário de tipos
    df_tipos = pd.DataFrame([{"Código": k, "Descrição": v} for k, v in TIPO_PERGUNTA_CHOICES])

    # Alternativas
    alternativas = Alternativa.objects.select_related("pergunta")
    alt_data = [{
        "Fase": alt.pergunta.fase_tratamento,
        "Número da Pergunta": alt.pergunta.numero_pergunta,
        "Pergunta": alt.pergunta.texto,
        "Alternativa": alt.texto
    } for alt in alternativas]
    df_alternativas = pd.DataFrame(alt_data)

    # Exportar para Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=dados_bioestatistica.xlsx"

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df_estatisticas.to_excel(writer, index=False, sheet_name="Estatísticas")
        df_brutos.to_excel(writer, index=False, sheet_name="Dados Brutos")
        df_fases.to_excel(writer, index=False, sheet_name="Dicionário Fases")
        df_tipos.to_excel(writer, index=False, sheet_name="Dicionário Tipos")
        df_alternativas.to_excel(writer, index=False, sheet_name="Alternativas")

        # Formatação
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                worksheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    return response